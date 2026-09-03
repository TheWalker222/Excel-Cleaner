import re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from datetime import datetime
moneyPattern = re.compile(r"^(?:[$€₹]\s*[\d.,\s]+|[\d.,\s]+\s*[$€₹])$")
datePattern = re.compile(r"\d{1,2}[\.\/\-]\d{1,2}[\.\/\-]\d{2}\b|\b\d{1,2}[\.\/\-]\d{1,2}[\.\/\-]\d{4}")
zeroPattern = re.compile(r"^0+$|^0+[1-9]")
urlPattern = re.compile(r"(?:https://[^\s,]*)\b|(?:http://[^\s,]*)\b|(?:[^\s,]*\.(?:com|org|fi|ru|me)(?:/[^\s,]*)*)\b")
emailPattern = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
phoneNumberPattern = re.compile(r"(?:\+?\d{1,3})(?:[\s-]?\(?\d{2,3}\)?){2,4}\b|0(?:[\s-]?\(?\d{2,3}\)?){2,4}\b")
bad_patterns = {
    "..": ".",
    ",,": ",",
    "@@": "@",
    "__": "_",
    "--": "-",
}
def check_bad_patterns(value):
    if any(pattern in value for pattern in bad_patterns):
        for pattern, replacement in bad_patterns.items():
            value = value.replace(pattern, replacement)
        return check_bad_patterns(value)
    else:
        return value
def check_value(value):
    value = str(value).strip()
    value = check_bad_patterns(value)
    match = urlPattern.fullmatch(value)
    if match:
        return value, "url"
    match = emailPattern.fullmatch(value)
    if match:
        return value, "email"
    match = phoneNumberPattern.fullmatch(value)
    if match:
        return value, "phone"
    match = moneyPattern.fullmatch(value)
    if match:
        value = value.replace("$", "").replace("€", "").replace("₹", "")
        value = value.replace(" ", "")
        if "," in value and "." in value:
            if value.rfind(".") > value.rfind(","):
                value = value.replace(",", "")
            else:
                value = value.replace(".", "").replace(",", ".")
        elif "," in value:
            right = value.split(",")[-1]
            if len(right) == 3:
                value = value.replace(",", "")
            else:
                value = value.replace(",", ".")
        elif "." in value:
            right = value.split(".")[-1]
            if len(right) == 3:
                value = value.replace(".", "")
        return value, "money"
    match = datePattern.fullmatch(value)
    if match:
        for fmt in ("%d.%m.%y","%d.%m.%Y","%d/%m/%y","%d/%m/%Y","%d-%m-%y","%d-%m-%Y"):
            try:
                value = datetime.strptime(value, fmt)
                break
            except ValueError:
                continue
        return value, "date"
    match = zeroPattern.fullmatch(value)
    if match:
        return value, "text"
    if value.isdigit():
        return value, "int"
    try:
        value = value.replace(" ", "")
        if "," in value and "." in value:
            if value.rfind(".") > value.rfind(","):
                value = value.replace(",", "")
            else:
                value = value.replace(".", "").replace(",", ".")
        elif "," in value:
            right = value.split(",")[-1]
            if len(right) == 3:
                value = value.replace(",", "")
            else:
                value = value.replace(",", ".")
        elif "." in value:
            right = value.split(".")[-1]
            if len(right) == 3:
                value = value.replace(".", "")
        float(value)
        return value, "float"
    except ValueError:
        return value, "text"
def formatSheet(output_ws):
    for column in output_ws.columns:
        max_length = 0
        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        output_ws.column_dimensions[column[0].column_letter].width = adjusted_width
def check_exceptions(value):
    value = str(value).lower().strip()
    words = value.split()
    exceptions = {
        "mc'donalds": "McDonald's",
        "usa": "USA",
        "uk": "UK",
        "u.s.a": "USA",
        "u.k": "UK",
        "u.s": "USA",
        "u.k.": "UK",
        "us": "US",
        "iphone": "iPhone",
        "ipad": "iPad",
        "macbook": "MacBook",
        "api": "API",
    }
    words = [exceptions.get(word, word.title()) for word in words]
    return " ".join(words)
def remove_empty_rows(ws):
    rows_to_delete = []
    for row in ws.iter_rows():
        if all(cell.value is None or str(cell.value).strip() == "" for cell in row):
            rows_to_delete.append(row[0].row)
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)
def remove_duplicate_rows(ws):
    seen_rows = set()
    rows_to_delete = []
    for row in ws.iter_rows():
        values = tuple(cell.value for cell in row)
        if values in seen_rows:
            rows_to_delete.append(row[0].row)
        else:
            seen_rows.add(values)
    for row in reversed(rows_to_delete):
        ws.delete_rows(row)
def clean_excel_file(files, delete_duplicates, delete_empty):
    for file in files:
        input_wb = load_workbook(file)
        output_wb = Workbook()
        output_wb.remove(output_wb.active)
        for ws in input_wb.worksheets:
            output_ws = output_wb.create_sheet(title=ws.title)
            for rowI, row in enumerate(ws.rows, start=1):
                for columnI, cell in enumerate(row, start=1):
                    value = cell.value
                    if value == "" or value == "None" or value == "Nan" or value == "Na" or value == "N/A" or value is None:
                        output_ws.cell(row=rowI, column=columnI).value = ""
                        continue
                    value,method = check_value(value)
                    if method == "money":
                        excelCell = output_ws.cell(row=rowI, column=columnI)
                        excelCell.value = float(value)
                    elif method == "date":
                        excelCell = output_ws.cell(row=rowI, column=columnI)
                        excelCell.value = value
                        excelCell.number_format = 'DD.MM.YYYY'
                    elif method == "int":
                        excelCell = output_ws.cell(row=rowI, column=columnI)
                        excelCell.value = int(value)
                    elif method == "float":
                        excelCell = output_ws.cell(row=rowI, column=columnI)
                        excelCell.value = float(value)
                    elif method == "url":
                        excelCell = output_ws.cell(row=rowI, column=columnI)
                        excelCell.value = value
                    elif method == "email":
                        excelCell = output_ws.cell(row=rowI, column=columnI)
                        excelCell.value = value.lower()
                    elif method == "phone":
                        excelCell = output_ws.cell(row=rowI, column=columnI)
                        excelCell.value = value
                    else:
                        value = check_exceptions(value)
                        excelCell = output_ws.cell(row=rowI, column=columnI)
                        excelCell.value = value
                if rowI == 1:
                    for cell in output_ws[1]:
                        cell.font = Font(bold=True)
            output_ws.auto_filter.ref = output_ws.dimensions
            output_ws.freeze_panes = output_ws['A2']
            if delete_empty:
                remove_empty_rows(output_ws)
            if delete_duplicates:
                remove_duplicate_rows(output_ws)
            formatSheet(output_ws)
        output_wb.save(file)